"""Exportació de contractes: job, abast, CSV/XLSX i descàrrega amb token."""

import io
from collections.abc import AsyncIterator
from typing import Any
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine
from sqlalchemy.pool import NullPool

from app.core.config import settings
from app.core.db import session_factory
from app.jobs.registry import JobContext
from app.modules.contracts.exports import export_contracts
from tests.conftest import login_headers


@pytest.fixture
async def world(  # type: ignore[no-untyped-def]
    make_user, monkeypatch: pytest.MonkeyPatch, tmp_path
) -> AsyncIterator[dict[str, Any]]:
    monkeypatch.setattr(settings, "storage_backend", "filesystem")
    monkeypatch.setattr(settings, "storage_local_path", str(tmp_path))

    tag = uuid4().hex[:8]
    data: dict[str, Any] = {"tag": tag, "storage": tmp_path}
    data["admin"] = await make_user("admin")
    data["employee"] = await make_user("employee")

    async with session_factory() as session:
        dept = (
            await session.execute(
                text("INSERT INTO departments (code, name) VALUES (:c, 'EXP') RETURNING id"),
                {"c": f"EX-{tag}"},
            )
        ).scalar_one()
        await session.execute(
            text("INSERT INTO user_departments (user_id, department_id) VALUES (:u, :d)"),
            {"u": data["employee"].id, "d": dept},
        )
        in_dept = (
            await session.execute(
                text(
                    "INSERT INTO contracts (file_code, status, lot, subject, award_amount) "
                    "VALUES (:f, 'Execució', '', :s, 1234.56) RETURNING id"
                ),
                {"f": f"EXP-{tag}/1", "s": f"Objecte; amb punt i coma {tag}"},
            )
        ).scalar_one()
        await session.execute(
            text("INSERT INTO contract_departments (contract_id, department_id) VALUES (:c, :d)"),
            {"c": in_dept, "d": dept},
        )
        await session.execute(
            text(
                "INSERT INTO contracts (file_code, status, lot, subject) "
                "VALUES (:f, 'Execució', '', :s)"
            ),
            {"f": f"EXP-{tag}/2", "s": f"Fora d'abast {tag}"},
        )
        await session.commit()
        data.update(dept=dept, in_dept=in_dept)

    yield data

    engine = create_async_engine(settings.database_url, poolclass=NullPool)
    async with engine.begin() as conn:
        await conn.execute(
            text("DELETE FROM contracts WHERE file_code LIKE :p"), {"p": f"EXP-{tag}%"}
        )
        await conn.execute(text("DELETE FROM departments WHERE code LIKE :p"), {"p": f"EX-{tag}"})
    await engine.dispose()


async def _run_export(payload: dict[str, Any]) -> dict[str, Any]:
    async def _noop(_pct: int, _msg: str | None = None) -> None:
        return None

    result = await export_contracts(JobContext(job_id=uuid4(), payload=payload, set_progress=_noop))
    assert result is not None
    return result


async def test_csv_respects_scope_and_format(world: dict[str, Any]) -> None:
    tag = world["tag"]
    result = await _run_export(
        {
            "format": "csv",
            "user_id": world["employee"].id,
            "scope": {"type": "departments", "department_ids": [world["dept"]]},
            "filters": {"q": tag},
        }
    )
    assert result["rows"] == 1  # el contracte fora d'abast NO hi és

    content = (world["storage"] / result["storage_key"]).read_bytes()
    assert content.startswith(b"\xef\xbb\xbf")  # BOM
    text_content = content.decode("utf-8-sig")
    lines = text_content.strip().splitlines()
    assert lines[0].startswith("Expedient;Lot;Estat")
    assert f"EXP-{tag}/1" in lines[1]
    # El punt i coma de l'objecte queda entre cometes (CSV vàlid).
    assert '"Objecte; amb punt i coma' in lines[1]
    assert "1234.56" in lines[1]


async def test_xlsx_contains_same_rows(world: dict[str, Any]) -> None:
    from openpyxl import load_workbook

    tag = world["tag"]
    result = await _run_export(
        {
            "format": "xlsx",
            "user_id": world["admin"].id,
            "scope": {"type": "all"},
            "filters": {"q": tag},
        }
    )
    assert result["rows"] == 2  # admin ho veu tot

    content = (world["storage"] / result["storage_key"]).read_bytes()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet.max_row == 3  # capçalera + 2 files
    assert sheet["A1"].value == "Expedient"


async def test_export_endpoint_and_single_use_download(
    api_client: TestClient, world: dict[str, Any]
) -> None:
    tag = world["tag"]
    headers = login_headers(api_client, world["employee"].email)

    # Encuament (l'employee pot exportar el seu abast).
    response = api_client.post(
        "/api/v1/contracts/exports",
        json={"format": "csv", "filters": {"q": tag}},
        headers=headers,
    )
    assert response.status_code == 202, response.text
    job_id = response.json()["id"]

    # view=all sense dret: 403.
    denied = api_client.post(
        "/api/v1/contracts/exports",
        json={"format": "csv", "view": "all"},
        headers=headers,
    )
    assert denied.status_code == 403

    # S'executa el job directament (sense worker als tests).
    from uuid import UUID

    from app.jobs.models import Job, JobStatus

    result = await _run_export(
        {
            "format": "csv",
            "user_id": world["employee"].id,
            "scope": {"type": "departments", "department_ids": [world["dept"]]},
            "filters": {"q": tag},
        }
    )
    async with session_factory() as session:
        job = await session.get(Job, UUID(job_id))
        assert job is not None
        job.status = JobStatus.SUCCESS
        job.result = result
        await session.commit()

    # Token efímer amb propòsit download.
    token_response = api_client.post(
        "/api/v1/auth/ephemeral",
        json={"purpose": "download", "resource": job_id},
        headers=headers,
    )
    assert token_response.status_code == 201, token_response.text
    token = token_response.json()["token"]

    # Descàrrega sense capçalera d'autenticació: el token és l'autorització.
    download = api_client.get(f"/api/v1/contracts/exports/{job_id}/download?token={token}")
    assert download.status_code == 200, download.text
    assert download.headers["content-type"].startswith("text/csv")
    assert "attachment" in download.headers["content-disposition"]
    assert f"EXP-{tag}/1" in download.text

    # Un sol ús: la segona descàrrega falla.
    second = api_client.get(f"/api/v1/contracts/exports/{job_id}/download?token={token}")
    assert second.status_code == 401

    # Token d'un recurs diferent: 401.
    other = api_client.post(
        "/api/v1/auth/ephemeral",
        json={"purpose": "job_events", "resource": job_id},
        headers=headers,
    )
    wrong_purpose = api_client.get(
        f"/api/v1/contracts/exports/{job_id}/download?token={other.json()['token']}"
    )
    assert wrong_purpose.status_code == 401

    async with session_factory() as session:
        await session.execute(text("DELETE FROM jobs WHERE id = :id"), {"id": job_id})
        await session.commit()
